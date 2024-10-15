import camelot
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Constants
MONTH_REGEX = r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\b'
HIGHLIGHT_COLUMNS = ["Partial Charges ($)", "Monthly and Other Charges ($)", "Add-Ons ($)", "Usage Charges ($)"]

# Utility Functions
def format_contact_number(contact_num):
    return re.sub(r'(\d{3})[ ](\d{3}-\d{4})', r'\1-\2', contact_num)

def dash_to_zero(value):
    return '0' if value == '-' else value

def convert_to_number(value):
    value = dash_to_zero(value)
    try:
        return float(value.replace(',', '').replace('$', '').strip())
    except ValueError:
        return 0

def parse_pages_input(page_input):
    pages = []
    for part in page_input.split(','):
        part = part.strip()
        if '-' in part:
            start, end = map(int, part.split('-'))
            pages.extend(range(start, end + 1))
        else:
            pages.append(int(part))
    return pages

def adjust_column_widths(worksheet):
    for column_cells in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

def get_headers(df, table_type):
    if table_type.lower() == 'hardware':
        return ['First Name', 'Last Name', 'Contact Number',
                'Starting Balance', 'Payments ($)', 'Current Balance',
                'Starting Device Discount Balance ($)', 'Current Device Discount Balance ($)']
    else:
        return (['First Name', 'Last Name', 'Contact Number', 'Partial Charges ($)',
                 'Monthly and Other Charges ($)', 'Add-Ons ($)', 'Usage Charges ($)',
                 'Total Before Taxes ($)', 'Taxes ($)', 'Total ($)'] if len(df.columns) == 8
                else ['First Name', 'Last Name', 'Contact Number',
                      'Monthly and Other Charges ($)', 'Add-Ons ($)', 'Usage Charges ($)',
                      'Total Before Taxes ($)', 'Taxes ($)', 'Total ($)'])

def clean_data_row(row, contact_num, table_type, df_columns_count):
    first_name, last_name = row[0].split(maxsplit=1)
    
    if table_type.lower() == 'hardware':
        return [
            first_name, last_name, contact_num,
            convert_to_number(row[1].strip()), convert_to_number(row[2].strip()),
            convert_to_number(row[3].strip()), convert_to_number(row[4].strip()),
            convert_to_number(row[5].strip())
        ]
    elif df_columns_count == 8:
        return [
            first_name, last_name, contact_num,
            convert_to_number(row[1].strip()), convert_to_number(row[2].strip()),
            convert_to_number(row[3].strip()), convert_to_number(row[4].strip()),
            convert_to_number(row[5].strip()), convert_to_number(row[6].strip()),
            convert_to_number(row[7].strip())
        ]
    else:
        return [
            first_name, last_name, contact_num,
            convert_to_number(row[1].strip()), convert_to_number(row[2].strip()),
            convert_to_number(row[3].strip()), convert_to_number(row[4].strip()),
            convert_to_number(row[5].strip()), convert_to_number(row[6].strip())
        ]

def clean_data_frame(df, skip_conditions, table_type):
    headers = get_headers(df, table_type)
    cleaned_data = []

    i = 0
    while i < len(df):
        row = df.iloc[i]
        first_column_text = row[0].strip()

        if re.search(MONTH_REGEX, first_column_text):
            i += 1
            continue
        if "summary of mobile data sharing" in first_column_text.lower():
            break
        if any(skip_text.lower() in first_column_text.lower() for skip_text in skip_conditions):
            i += 1
            continue

        if len(first_column_text.split()) > 1:
            contact_row = df.iloc[i + 1] if i + 1 < len(df) else None
            contact_num = format_contact_number(contact_row[0].strip()) if contact_row is not None else ''

            cleaned_data.append(clean_data_row(row, contact_num, table_type, len(df.columns)))
            i += 2
        else:
            i += 1

    return pd.DataFrame(cleaned_data, columns=headers)

def highlight_columns(worksheet, headers):
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for col in worksheet.iter_cols(min_row=1, max_row=1):
        if col[0].value in HIGHLIGHT_COLUMNS:
            for cell in col:
                cell.fill = highlight_fill

def save_to_excel(output_path, final_data, special_page_data):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        combined_df = pd.concat(final_data, ignore_index=True)
        combined_df.to_excel(writer, index=False, header=True, sheet_name="Extracted Data")

        if special_page_data is not None:
            special_page_data.to_excel(writer, index=False, header=True, sheet_name="Review")
            highlight_columns(writer.sheets["Review"], special_page_data.columns)

def run_extraction(pdf_path, table_type, page_input):
    try:
        pages = parse_pages_input(page_input)
        tables = camelot.read_pdf(pdf_path, flavor='stream', pages=",".join(map(str, pages)))

        skip_conditions = ["SAMSUNG", "IPHONE", "GOOGLE", "GALAXY", "SUMMARY", "MOBILE", "SPAAR"] if table_type.lower() == 'hardware' \
            else ["BBAN", "BUSINESS", "MOBILE", "SUMMARY", "ACCOUNT", "TABLET", "SPAAR"]

        final_data = []
        special_page_data = None

        for page_number, table in zip(pages, tables):
            df = table.df
            cleaned_df = clean_data_frame(df, skip_conditions, table_type)

            if page_number == 26:
                special_page_data = cleaned_df
            else:
                final_data.append(cleaned_df)

        return final_data, special_page_data

    except Exception as e:
        raise e
