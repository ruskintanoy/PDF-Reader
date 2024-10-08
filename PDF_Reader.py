import camelot
import pandas as pd
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

MONTH_REGEX = r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\b'
HIGHLIGHT_COLUMNS = ["Partial Charges ($)", "Monthly and Other Charges ($)", "Add-Ons ($)", "Usage Charges ($)"]

def format_contact_number(contact_num):
    return re.sub(r'(\d{3})[ ](\d{3}-\d{4})', r'\1-\2', contact_num)

def dash_to_zero(value):
    return '0' if value == '-' else value

def convert_to_number(value):
    value = dash_to_zero(value)
    try:
        return float(value.replace(',', '').replace('$', '').strip())
    except ValueError:
        return 0

def parse_pages_input(page_input):
    pages = []
    for part in page_input.split(','):
        part = part.strip()
        if '-' in part:
            start, end = map(int, part.split('-'))
            pages.extend(range(start, end + 1))
        else:
            pages.append(int(part))
    return pages

def adjust_column_widths(worksheet):
    for column_cells in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

def get_headers(df, table_type):
    if table_type.lower() == 'hardware':
        return ['First Name', 'Last Name', 'Contact Number',
                'Starting Balance', 'Payments ($)', 'Current Balance',
                'Starting Device Discount Balance ($)', 'Current Device Discount Balance ($)']
    else:
        return (['First Name', 'Last Name', 'Contact Number', 'Partial Charges ($)',
                 'Monthly and Other Charges ($)', 'Add-Ons ($)', 'Usage Charges ($)',
                 'Total Before Taxes ($)', 'Taxes ($)', 'Total ($)'] if len(df.columns) == 8
                else ['First Name', 'Last Name', 'Contact Number',
                      'Monthly and Other Charges ($)', 'Add-Ons ($)', 'Usage Charges ($)',
                      'Total Before Taxes ($)', 'Taxes ($)', 'Total ($)'])

def clean_data_frame(df, skip_conditions, table_type):
    headers = get_headers(df, table_type)
    cleaned_data = []

    i = 0
    while i < len(df):
        row = df.iloc[i]
        first_column_text = row[0].strip()

        if re.search(MONTH_REGEX, first_column_text):
            i += 1
            continue
        if "summary of mobile data sharing" in first_column_text.lower():
            break
        if any(skip_text.lower() in first_column_text.lower() for skip_text in skip_conditions):
            i += 1
            continue

        if len(first_column_text.split()) > 1:
            name_parts = first_column_text.split(maxsplit=1)
            first_name, last_name = name_parts[0], name_parts[1] if len(name_parts) > 1 else ''
            contact_row = df.iloc[i + 1] if i + 1 < len(df) else None
            contact_num = format_contact_number(contact_row[0].strip()) if contact_row is not None else ''

            if table_type.lower() == 'hardware':
                cleaned_data.append([
                    first_name, last_name, contact_num,
                    convert_to_number(row[1].strip()), convert_to_number(row[2].strip()),
                    convert_to_number(row[3].strip()), convert_to_number(row[4].strip()),
                    convert_to_number(row[5].strip())
                ])
            else:
                if len(df.columns) == 8:
                    cleaned_data.append([
                        first_name, last_name, contact_num,
                        convert_to_number(row[1].strip()), convert_to_number(row[2].strip()),
                        convert_to_number(row[3].strip()), convert_to_number(row[4].strip()),
                        convert_to_number(row[5].strip()), convert_to_number(row[6].strip()),
                        convert_to_number(row[7].strip())
                    ])
                else:
                    cleaned_data.append([
                        first_name, last_name, contact_num,
                        convert_to_number(row[1].strip()), convert_to_number(row[2].strip()),
                        convert_to_number(row[3].strip()), convert_to_number(row[4].strip()),
                        convert_to_number(row[5].strip()), convert_to_number(row[6].strip())
                    ])
            i += 2
        else:
            i += 1

    return pd.DataFrame(cleaned_data, columns=headers)

def highlight_columns(worksheet, headers):
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for col in worksheet.iter_cols(min_row=1, max_row=1):
        if col[0].value in HIGHLIGHT_COLUMNS:
            for cell in col:
                cell.fill = highlight_fill

def run_extraction(pdf_path, table_type, page_input, root):
    try:
        pages = parse_pages_input(page_input)
        tables = camelot.read_pdf(pdf_path, flavor='stream', pages=",".join(map(str, pages)))
        final_data = []
        page_26_data = None

        if table_type.lower() == 'hardware':
            skip_conditions = ["SAMSUNG", "IPHONE", "GOOGLE", "GALAXY", "SUMMARY", "MOBILE", "SPAAR"]
        else:
            skip_conditions = ["BBAN", "BUSINESS", "MOBILE", "SUMMARY", "ACCOUNT", "TABLET", "SPAAR"]

        for page_number, table in zip(pages, tables):
            df = table.df
            cleaned_df = clean_data_frame(df, skip_conditions, table_type)

            if page_number == 26:
                page_26_data = cleaned_df
            else:
                final_data.append(cleaned_df)

        output_excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        if not output_excel_path:
            messagebox.showerror("Save Error", "No save location selected.")
            return

        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
         
            combined_df = pd.concat(final_data, ignore_index=True)
            combined_df.to_excel(writer, index=False, header=True, sheet_name="Extracted Data")

            if page_26_data is not None:
                page_26_data.to_excel(writer, index=False, header=True, sheet_name="Review")
                highlight_columns(writer.sheets["Review"], page_26_data.columns)

        workbook = load_workbook(output_excel_path)

        worksheet_data = workbook["Extracted Data"]
        adjust_column_widths(worksheet_data)

        if "Review" in workbook.sheetnames:
            worksheet_review = workbook["Review"]
            adjust_column_widths(worksheet_review)

        workbook.save(output_excel_path)

        root.quit()
        root.destroy()

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def open_ui():
    root = tk.Tk()
    root.title("PDF Table Extractor")
    root.geometry("400x300")

    def browse_pdf():
        pdf_path.set(filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")]))

    def extract():
        pdf = pdf_path.get()
        table_type = table_type_var.get()
        pages = page_entry.get()

        if not pdf or not table_type or not pages:
            messagebox.showerror("Input Error", "All fields are required!")
        else:
            run_extraction(pdf, table_type, pages, root)

    pdf_path = tk.StringVar()
    table_type_var = tk.StringVar()

    tk.Label(root, text="Select PDF File:").pack(pady=5)
    tk.Entry(root, textvariable=pdf_path, width=50).pack(pady=5)
    tk.Button(root, text="Browse", command=browse_pdf).pack(pady=5)

    tk.Label(root, text="Table Type:").pack(pady=5)
    table_type_dropdown = Combobox(root, textvariable=table_type_var)
    table_type_dropdown['values'] = ('Hardware', 'Raw')
    table_type_dropdown.pack(pady=5)

    tk.Label(root, text="Pages (e.g., 1,2,5-7):").pack(pady=5)
    page_entry = tk.Entry(root, width=20)
    page_entry.pack(pady=5)

    tk.Button(root, text="Extract", command=extract, bg="green", fg="white").pack(pady=20)

    root.mainloop()

open_ui()