import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from openpyxl import load_workbook
from reader import run_extraction, save_to_excel, adjust_column_widths

# UI Logic
def open_ui():
    root = tk.Tk()
    root.title("PDF Table Extractor")
    root.geometry("400x300")

    pdf_path = tk.StringVar()
    table_type_var = tk.StringVar()

    def browse_pdf():
        pdf_path.set(filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")]))

    def extract():
        pdf = pdf_path.get()
        table_type = table_type_var.get()
        pages = page_entry.get()

        if not pdf or not table_type or not pages:
            messagebox.showerror("Input Error", "All fields are required!")
        else:
            try:
                final_data, special_page_data = run_extraction(pdf, table_type, pages)

                output_excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
                if not output_excel_path:
                    messagebox.showerror("Save Error", "No save location selected.")
                    return

                save_to_excel(output_excel_path, final_data, special_page_data)

                workbook = load_workbook(output_excel_path)
                adjust_column_widths(workbook["Extracted Data"])

                if "Review" in workbook.sheetnames:
                    adjust_column_widths(workbook["Review"])

                workbook.save(output_excel_path)

                root.quit()
                root.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {e}")

    tk.Label(root, text="Select PDF File:").pack(pady=5)
    tk.Entry(root, textvariable=pdf_path, width=50).pack(pady=5)
    tk.Button(root, text="Browse", command=browse_pdf).pack(pady=5)

    tk.Label(root, text="Table Type:").pack(pady=5)
    table_type_dropdown = Combobox(root, textvariable=table_type_var)
    table_type_dropdown['values'] = ('Hardware', 'Raw')
    table_type_dropdown.pack(pady=5)

    tk.Label(root, text="Pages (e.g., 1,2,5-7):").pack(pady=5)
    page_entry = tk.Entry(root, width=20)
    page_entry.pack(pady=5)

    tk.Button(root, text="Extract", command=extract, bg="green", fg="white").pack(pady=20)

    root.mainloop()

open_ui()
